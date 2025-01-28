import csv
import tkinter as tk
from tkinter import messagebox, ttk
from ttkbootstrap import Style
import openpyxl
import datetime

# Charger les données du quiz depuis le fichier CSV
def load_quiz_data_from_csv(filename):
    quiz_data = []
    try:
        with open(filename, "r", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                quiz_data.append({
                    "question": row["Question"],
                    "choices": [row["Choice 1"], row["Choice 2"], row["Choice 3"], row["Choice 4"]],
                    "answer": row["Correct Answer"]
                })
    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible de charger le fichier CSV : {e}")
        root.destroy()
    return quiz_data

# Charger le quiz_data
quiz_data = load_quiz_data_from_csv("quiz_questions.csv")
if not quiz_data:
    messagebox.showerror("Erreur", "Le fichier CSV est vide ou mal formaté.")
    exit()

# Liste des concepts associés aux questions
concepts = [
    "Modèle OSI",
    "Modèle OSI",
    "Ethernet",
    "Ethernet",
    "Protocole IP",
    "Protocole IP",
    "Switch niveau 2",
    "Switch niveau 2",
    "Câbles Ethernets",
    "Câbles Ethernets",
    "Fibre optique",
    "Fibre optique",
    "Cartes réseau modernes",
    "Cartes réseau modernes",
    "Agrégation de liens",
    "Agrégation de liens"
]

# Liste des concepts détaillés
detailed_concepts = [
    "Couches OSI",
    "Fonctionnalités des couches",
    "Câbles Ethernet",
    "Normes IEEE 802.3",
    "TCP",
    "UDP",
    "VLAN",
    "Spanning Tree Protocol",
    "STP",
    "UTP",
    "Fibre multimode",
    "Fibre monomode",
    "Optimisation réseau",
    "Adresses MAC",
    "Augmentation bande passante",
    "Redondance"
]

# Fonction pour sauvegarder les résultats dans un fichier Excel spécifique à l'étudiant
def save_results_to_excel():
    # Générer le nom du fichier en fonction du nom de l'étudiant
    filename = f"quiz_results_{user_name}.xlsx"

    # Charger ou créer le fichier Excel
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Générer le nom de la feuille
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")
    sheet_name = f"Results-{timestamp}"

    # Vérifier si le nom de la feuille est unique (OpenPyxl limite à 31 caractères pour les noms de feuille)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    # Créer une nouvelle feuille
    sheet = workbook.create_sheet(title=sheet_name)

    # Écrire les en-têtes
    headers = ["Name", "Concept", "Detailed Concept", "Question", "Correct Answer", "Your Answer", "Result", "Score"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Écrire les résultats avec les concepts et les concepts détaillés
    for row_num, result in enumerate(results, 2):
        concept_index = (row_num - 2) % len(concepts)  # Associer chaque question à un concept
        detailed_concept_index = (row_num - 2) % len(detailed_concepts)  # Associer chaque question à un concept détaillé
        sheet.cell(row=row_num, column=1, value=user_name)
        sheet.cell(row=row_num, column=2, value=concepts[concept_index])  # Ajouter le concept
        sheet.cell(row=row_num, column=3, value=detailed_concepts[detailed_concept_index])  # Ajouter le concept détaillé
        sheet.cell(row=row_num, column=4, value=result["Question"])
        sheet.cell(row=row_num, column=5, value=result["Correct Answer"])
        sheet.cell(row=row_num, column=6, value=result["Your Answer"])
        sheet.cell(row=row_num, column=7, value=result["Result"])
        # Ajouter le score : 1 pour correct, 0 pour incorrect
        score_value = 1 if result["Result"] == "Correct" else 0
        sheet.cell(row=row_num, column=8, value=score_value)

    # Supprimer la feuille par défaut si elle est encore vide
    default_sheet = workbook["Sheet"] if "Sheet" in workbook.sheetnames else None
    if default_sheet and default_sheet.max_row == 1:
        workbook.remove(default_sheet)

    # Sauvegarder le fichier Excel sous le nom spécifique
    workbook.save(filename)

# Fonction pour afficher la question actuelle et ses choix
def show_question():
    question = quiz_data[current_question]
    qs_label.config(text=question["question"])
    question_number_label.config(text=f"Question {current_question + 1}/{len(quiz_data)}")

    choices = question["choices"]
    for i in range(4):
        choice_btns[i].config(text=choices[i], state="normal", style="TButton")

    feedback_label.config(text="")
    correct_answer_label.config(text="")
    next_btn.config(state="disabled")

# Fonction pour vérifier la réponse sélectionnée
def check_answer(choice):
    global results, score
    question = quiz_data[current_question]
    selected_choice = choice_btns[choice].cget("text").strip().lower()
    correct_answer = question["answer"].strip().lower()

    # Réinitialiser les styles des boutons
    for button in choice_btns:
        button.config(style="TButton")

    # Vérification de la réponse
    if selected_choice == correct_answer:
        score += 1
        choice_btns[choice].config(style="success.TButton")
        feedback_label.config(text="Correct!", foreground="green")
    else:
        choice_btns[choice].config(style="danger.TButton")
        for i, button in enumerate(choice_btns):
            if button.cget("text").strip().lower() == correct_answer:
                button.config(style="success.TButton")
        feedback_label.config(text="Incorrect!", foreground="red")
        correct_answer_label.config(text=f"Correct answer: {question['answer']}", foreground="green")

    score_label.config(text=f"Score: {score}/{len(quiz_data)}")

    results.append({
        "Question": question["question"],
        "Correct Answer": question["answer"],
        "Your Answer": choice_btns[choice].cget("text"),
        "Result": "Correct" if selected_choice == correct_answer else "Incorrect"
    })

    for button in choice_btns:
        button.config(state="disabled")
    next_btn.config(state="normal")

# Fonction pour passer à la question suivante
def next_question():
    global current_question
    current_question += 1

    if current_question < len(quiz_data):
        show_question()
    else:
        save_results_to_excel()
        messagebox.showinfo(
            "Quiz Completed",
            f"Quiz Completed! Final score: {score}/{len(quiz_data)}\nResults saved to 'quiz_results_{user_name}.xlsx'."
        )
        root.destroy()

# Fonction pour démarrer le quiz après saisie du nom
def start_quiz():
    global user_name
    user_name = name_entry.get().strip()
    if not user_name:
        messagebox.showerror("Error", "Please enter your name to start the quiz.")
        return

    name_frame.destroy()
    quiz_frame.pack()
    show_question()

# Fonction pour quitter le quiz
def quit_quiz():
    quit_confirmation = messagebox.askyesno("Quit", "Are you sure you want to quit?")
    if quit_confirmation:
        root.destroy()

# Création de la fenêtre principale
root = tk.Tk()
root.title("Quiz App")

# Plein écran
root.attributes("-fullscreen", True)

style = Style(theme="flatly")
style.configure("TLabel", font=("Helvetica", 20))
style.configure("TButton", font=("Helvetica", 16))
style.configure("success.TButton", background="green", foreground="white")
style.configure("danger.TButton", background="red", foreground="white")

# Frame pour la saisie du nom
name_frame = ttk.Frame(root)
name_frame.pack(pady=100, expand=True)

ttk.Label(name_frame, text="Enter your name:", font=("Helvetica", 16)).pack(pady=10)
name_entry = ttk.Entry(name_frame, font=("Helvetica", 16))
name_entry.pack(pady=10)
start_button = ttk.Button(name_frame, text="Start Quiz", command=start_quiz)
start_button.pack(pady=10)

# Frame pour le quiz
quiz_frame = ttk.Frame(root)
question_number_label = ttk.Label(quiz_frame, font=("Helvetica", 16), anchor="center", padding=10)
question_number_label.pack(pady=10)

qs_label = ttk.Label(quiz_frame, anchor="center", wraplength=600, justify="center", width=60, padding=20)
qs_label.pack(pady=20)

# Modifier la taille des boutons
choice_btns = []
for i in range(4):
    button = ttk.Button(quiz_frame, command=lambda i=i: check_answer(i), width=40)
    button.pack(pady=5, padx=20)
    choice_btns.append(button)

feedback_label = ttk.Label(quiz_frame, anchor="center", padding=10)
feedback_label.pack(pady=10)
correct_answer_label = ttk.Label(quiz_frame, anchor="center", padding=10)
correct_answer_label.pack(pady=10)

score = 0
score_label = ttk.Label(quiz_frame, text=f"Score: {score}/{len(quiz_data)}", anchor="center", padding=10)
score_label.pack(pady=10)

next_btn = ttk.Button(quiz_frame, text="Next", command=next_question, state="disabled")
next_btn.pack(pady=10)

quit_button = ttk.Button(root, text="✖ Quit", command=quit_quiz, style="danger")
quit_button.place(x=root.winfo_screenwidth() - 100, y=20, anchor="ne")

current_question = 0
results = []
user_name = ""

root.mainloop()